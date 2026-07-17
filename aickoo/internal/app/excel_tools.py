#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
@Project:    Aickoo-Assistant
@Author:     艾可科技（昆山）有限责任公司
@Contact:    aickoo@163.com
@CreateTime: 2026-07-07
@Copyright:  Copyright (c) 2026 艾可科技（昆山）有限责任公司
@License:    All Rights Reserved.
"""


import pandas as pd
import numpy as np
from typing import Optional, List, Dict, Any, Union, Callable
from pathlib import Path
import json
import warnings

warnings.filterwarnings('ignore')


class ExcelTool:
    """Excel操作工具 - 针对LLM使用场景优化"""

    # ==================== 读取核心 ====================
    @classmethod
    def read(
            cls,
            file_path: str,
            sheet_name: Optional[str] = None,
            rows_limit: int = 0,
            chunk_size: int = 50000,
            columns: Optional[List[str]] = None,
            sample_mode: str = "head_tail",  # head, tail, random, head_tail
            sample_size: int = 100,
            token_budget: int = 4000,
            engine: str = "openpyxl"
    ) -> Dict[str, Any]:
        """
        读取Excel文件，支持大文件优化

        Args:
            file_path: 文件路径
            sheet_name: 工作表名
            rows_limit: 最大读取行数(0=全部)
            chunk_size: 分块大小
            columns: 指定列
            sample_mode: 采样模式
            sample_size: 采样大小
            token_budget: Token预算
            engine: 读取引擎
        """
        path = Path(file_path)
        if not path.exists():
            return {"error": f"文件不存在: {file_path}"}

        # 1. 先检测文件元信息
        meta = cls._detect_metadata(file_path, sheet_name, engine)
        total_rows = meta.get("total_rows", 0)

        # 2. 判断是否需要分块处理
        if total_rows > chunk_size or rows_limit > chunk_size:
            return cls._read_large(
                file_path, sheet_name, rows_limit, chunk_size,
                columns, sample_mode, sample_size, token_budget, engine
            )
        else:
            return cls._read_small(
                file_path, sheet_name, rows_limit,
                columns, sample_mode, sample_size, token_budget, engine
            )

    # ==================== 小文件读取 ====================
    @classmethod
    def _read_small(cls, file_path, sheet_name, rows_limit,
                    columns, sample_mode, sample_size, token_budget, engine):
        """小文件直接读取"""
        df = pd.read_excel(
            file_path,
            sheet_name=sheet_name,
            nrows=rows_limit if rows_limit > 0 else None,
            usecols=columns if columns else None,
            engine=engine
        )

        result = cls._prepare_llm_response(df, sample_mode, sample_size, token_budget)
        result["meta"] = {
            "total_rows": len(df),
            "total_cols": len(df.columns),
            "memory_usage_mb": round(df.memory_usage(deep=True).sum() / 1024 / 1024, 2)
        }
        return result

    # ==================== 大文件读取（分块+采样） ====================
    @classmethod
    def _read_large(cls, file_path, sheet_name, rows_limit,
                    chunk_size, columns, sample_mode, sample_size,
                    token_budget, engine):
        """大文件分块读取 + 智能采样"""

        # 获取工作表名称
        if sheet_name is None:
            xl = pd.ExcelFile(file_path, engine=engine)
            sheet_name = xl.sheet_names[0]

        # 分块读取 + 采样收集
        chunks = []
        total_rows_read = 0

        # 使用openpyxl直接操作实现分块
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

        # 获取列头
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)

        # 确定要读取的列索引
        col_indices = None
        if columns:
            col_indices = [i for i, h in enumerate(headers) if h in columns]

        # 分块读取数据
        data_chunks = []
        current_chunk = []
        row_count = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if rows_limit > 0 and row_count >= rows_limit:
                break

            # 筛选列
            if col_indices:
                row_data = [row[i] for i in col_indices if i < len(row)]
            else:
                row_data = list(row)

            current_chunk.append(row_data)
            row_count += 1

            if len(current_chunk) >= chunk_size:
                # 保存当前块
                chunk_df = pd.DataFrame(current_chunk, columns=headers if not columns else columns)
                data_chunks.append(chunk_df)
                current_chunk = []

        # 处理最后一块
        if current_chunk:
            chunk_df = pd.DataFrame(current_chunk, columns=headers if not columns else columns)
            data_chunks.append(chunk_df)

        wb.close()

        # 合并采样
        if not data_chunks:
            return {"error": "未读取到数据"}

        # 根据采样模式从各块中采样
        sample_dfs = []
        total_rows = sum(len(c) for c in data_chunks)

        if sample_mode == "head":
            sample_dfs = [data_chunks[0].head(sample_size)]
        elif sample_mode == "tail":
            sample_dfs = [data_chunks[-1].tail(sample_size)]
        elif sample_mode == "random":
            # 从所有块中随机采样
            all_indices = []
            offset = 0
            for chunk in data_chunks:
                all_indices.extend(range(offset, offset + len(chunk)))
                offset += len(chunk)
            import random
            selected = random.sample(all_indices, min(sample_size, len(all_indices)))
            # 按块重新组织
            # 简化实现：从每块按比例采样
            per_chunk = max(1, sample_size // len(data_chunks))
            for chunk in data_chunks:
                if len(chunk) > per_chunk:
                    sample_dfs.append(chunk.sample(per_chunk))
                else:
                    sample_dfs.append(chunk)
        else:  # head_tail 默认
            sample_dfs = [data_chunks[0].head(sample_size // 2)]
            if len(data_chunks) > 1:
                sample_dfs.append(data_chunks[-1].tail(sample_size // 2))
            else:
                sample_dfs.append(data_chunks[0].tail(sample_size // 2))

        # 合并采样结果
        sample_df = pd.concat(sample_dfs).drop_duplicates().head(sample_size)

        # 计算总体统计（基于所有块）
        stats = cls._compute_chunked_stats(data_chunks)

        result = cls._prepare_llm_response(sample_df, "full", sample_size, token_budget)
        result["meta"] = {
            "total_rows": total_rows,
            "total_cols": len(data_chunks[0].columns) if data_chunks else 0,
            "chunks": len(data_chunks),
            "chunk_size": chunk_size,
            "is_large_file": True,
            "sample_mode": sample_mode,
            **stats
        }
        result["sample_info"] = {
            "sample_rows": len(sample_df),
            "sampled_from": total_rows
        }

        return result

    # ==================== 分块统计 ====================
    @classmethod
    def _compute_chunked_stats(cls, chunks: List[pd.DataFrame]) -> Dict:
        """计算分块数据的统计信息"""
        if not chunks:
            return {}

        numeric_cols = chunks[0].select_dtypes(include=[np.number]).columns.tolist()
        stats = {}

        for col in numeric_cols[:5]:  # 限制列数
            col_stats = {"min": float('inf'), "max": float('-inf'), "sum": 0, "count": 0}
            for chunk in chunks:
                if col in chunk.columns:
                    col_stats["min"] = min(col_stats["min"], chunk[col].min())
                    col_stats["max"] = max(col_stats["max"], chunk[col].max())
                    col_stats["sum"] += chunk[col].sum()
                    col_stats["count"] += len(chunk[col].dropna())
            if col_stats["count"] > 0:
                stats[col] = {
                    "min": col_stats["min"],
                    "max": col_stats["max"],
                    "mean": col_stats["sum"] / col_stats["count"],
                    "count": col_stats["count"]
                }

        return {"numeric_stats": stats}

    # ==================== LLM响应准备（核心：Token控制） ====================
    @classmethod
    def _prepare_llm_response(
            cls,
            df: pd.DataFrame,
            mode: str,
            sample_size: int,
            token_budget: int
    ) -> Dict[str, Any]:
        """
        准备LLM友好的响应，控制Token数量

        策略：
        1. 数据压缩为Markdown表格（LLM最易理解）
        2. 头部+尾部采样（保留上下文）
        3. 自动截断以适应Token预算
        """

        # 估算Token数（粗略：1 token ≈ 4字符）
        def estimate_tokens(text: str) -> int:
            return len(text) // 4

        result = {
            "data": None,
            "columns": df.columns.tolist(),
            "dtypes": df.dtypes.astype(str).to_dict(),
            "shape": df.shape,
            "compressed": True
        }

        # 1. 生成Markdown表格表示
        if len(df) > 0:
            # 限制显示行数
            display_df = df.head(sample_size)

            # 转换为Markdown
            md_table = cls._df_to_markdown(display_df)
            estimated_tokens = estimate_tokens(md_table)

            # 如果超出预算，进一步压缩
            if estimated_tokens > token_budget * 0.8:
                # 减少行数
                reduced_size = max(5, int(sample_size * token_budget / estimated_tokens))
                display_df = df.head(reduced_size)
                md_table = cls._df_to_markdown(display_df)

                # 如果还不够，减少列数
                if estimate_tokens(md_table) > token_budget * 0.8:
                    # 只保留前5列 + 关键列
                    key_cols = display_df.columns[:5].tolist()
                    display_df = display_df[key_cols]
                    md_table = cls._df_to_markdown(display_df)
                    result["truncated"] = True
                    result["truncated_reason"] = "Token budget exceeded, columns reduced"

            result["data"] = md_table
            result["token_estimate"] = estimate_tokens(md_table)

        # 2. 添加列统计（轻量级）
        result["column_summary"] = {}
        for col in df.columns[:10]:  # 限制列数
            col_info = {"type": str(df[col].dtype), "null_count": int(df[col].isna().sum())}
            if pd.api.types.is_numeric_dtype(df[col]):
                col_info.update({
                    "min": float(df[col].min()) if not df[col].isna().all() else None,
                    "max": float(df[col].max()) if not df[col].isna().all() else None,
                    "mean": float(df[col].mean()) if not df[col].isna().all() else None
                })
            elif pd.api.types.is_object_dtype(df[col]):
                unique_vals = df[col].dropna().unique()
                col_info["unique_count"] = len(unique_vals)
                col_info["sample_values"] = unique_vals[:5].tolist()
            result["column_summary"][col] = col_info

        return result

    @classmethod
    def _df_to_markdown(cls, df: pd.DataFrame, max_rows: int = 100) -> str:
        """将DataFrame转换为Markdown表格"""
        if len(df) == 0:
            return "*(空表格)*"

        # 限制行数
        display_df = df.head(max_rows)

        lines = []
        # 表头
        headers = display_df.columns.tolist()
        lines.append("| " + " | ".join(str(h) for h in headers) + " |")
        lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
        # 数据行
        for _, row in display_df.iterrows():
            cells = []
            for val in row:
                if pd.isna(val):
                    cells.append("")
                elif isinstance(val, (int, float)):
                    cells.append(str(val))
                else:
                    cells.append(str(val)[:50])  # 截断长文本
            lines.append("| " + " | ".join(cells) + " |")

        return "\n".join(lines)

    # ==================== 元数据检测 ====================
    @classmethod
    def _detect_metadata(cls, file_path: str, sheet_name: Optional[str],
                         engine: str) -> Dict:
        """检测文件元数据（不加载全部数据）"""
        try:
            # 使用openpyxl快速检测
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True, data_only=True)

            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.active

            total_rows = ws.max_row - 1  # 减去表头
            total_cols = ws.max_column

            wb.close()

            return {
                "total_rows": total_rows,
                "total_cols": total_cols,
                "sheet_name": ws.title
            }
        except Exception as e:
            return {"error": str(e), "total_rows": 0}

    # ==================== 摘要生成 ====================
    @classmethod
    def summarize(cls, file_path: str, sheet_name: Optional[str] = None,
                  token_budget: int = 3000) -> Dict[str, Any]:
        """
        生成数据摘要，帮助LLM快速理解数据
        """
        # 先读取小样本
        result = cls.read(
            file_path, sheet_name,
            rows_limit=1000,  # 只读1000行做摘要
            sample_mode="head_tail",
            sample_size=200,
            token_budget=token_budget
        )

        if "error" in result:
            return result

        # 增强摘要信息
        summary = {
            "file": file_path,
            "sheet": result.get("sheet_name", "unknown"),
            "shape": result.get("shape"),
            "columns": result.get("columns", []),
            "column_types": result.get("dtypes", {}),
            "sample_data": result.get("data", ""),
            "column_summary": result.get("column_summary", {}),
            "meta": result.get("meta", {})
        }

        # 生成自然语言摘要（供LLM直接使用）
        nl_summary = cls._generate_natural_language_summary(summary)
        summary["natural_language_summary"] = nl_summary

        return summary

    @classmethod
    def _generate_natural_language_summary(cls, summary: Dict) -> str:
        """生成自然语言摘要"""
        lines = []
        lines.append(f"📊 Excel数据摘要")
        lines.append(f"文件: {summary.get('file')}")
        lines.append(f"工作表: {summary.get('sheet')}")
        lines.append(f"数据规模: {summary.get('shape', ['?', '?'])[0]} 行 × {summary.get('shape', ['?', '?'])[1]} 列")

        lines.append(f"\n📋 列信息:")
        for col, info in summary.get('column_summary', {}).items():
            type_str = info.get('type', 'unknown')
            null_str = f"缺失{info.get('null_count', 0)}条"
            if 'min' in info:
                lines.append(
                    f"  - {col} ({type_str}): 范围 {info.get('min', '?')} ~ {info.get('max', '?')}, 均值 {info.get('mean', '?'):.2f}, {null_str}")
            elif 'unique_count' in info:
                samples = info.get('sample_values', [])[:3]
                lines.append(f"  - {col} ({type_str}): {info.get('unique_count', 0)}种取值, 样例 {samples}, {null_str}")
            else:
                lines.append(f"  - {col} ({type_str}): {null_str}")

        return "\n".join(lines)

    # ==================== 查询执行 ====================
    @classmethod
    def query(cls, file_path: str, query_expr: str, sheet_name: Optional[str] = None,
              limit: int = 1000) -> Dict[str, Any]:
        """
        执行Pandas查询

        Args:
            query_expr: 查询表达式，如 "df[df['age'] > 18]" 或 "df.groupby('city').size()"
        """
        try:
            # 读取数据（限制行数避免OOM）
            df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=100000)

            # 安全执行查询
            allowed_locals = {"df": df, "pd": pd, "np": np}
            result = eval(query_expr, {"__builtins__": {}}, allowed_locals)

            if isinstance(result, pd.DataFrame):
                if len(result) > limit:
                    result = result.head(limit)
                return {
                    "success": True,
                    "result": cls._df_to_markdown(result),
                    "shape": result.shape,
                    "columns": result.columns.tolist()
                }
            elif isinstance(result, pd.Series):
                return {
                    "success": True,
                    "result": result.to_dict(),
                    "type": "series"
                }
            else:
                return {
                    "success": True,
                    "result": str(result),
                    "type": "scalar"
                }
        except Exception as e:
            return {"success": False, "error": str(e)}

    # ==================== 新增：添加行 ====================
    @classmethod
    def add_row(
            cls,
            file_path: str,
            row_data: Dict[str, Any],
            sheet_name: Optional[str] = None,
            save: bool = False,
            output_path: Optional[str] = None,
            engine: str = "openpyxl"
    ) -> Dict[str, Any]:
        """
        向 Excel 添加一行数据。

        Args:
            row_data: 字典，键为列名，值为要添加的值
            save: 是否保存到文件
            output_path: 保存路径，若不指定则覆盖原文件
        """
        try:
            # 读取现有数据（限制 10 万行）
            df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=100000, engine=engine)
            if sheet_name is None:
                sheet_name = 0  # 默认第一个工作表

            # 添加行（使用 loc 或 append）
            new_row = pd.Series(row_data)
            df = pd.concat([df, new_row.to_frame().T], ignore_index=True)

            # 生成预览
            preview = cls._prepare_llm_response(df.tail(10), "head_tail", 10, 2000)
            preview["action"] = "add_row"
            preview["new_row"] = row_data
            preview["new_total_rows"] = len(df)

            if save:
                output_path = output_path or file_path
                df.to_excel(output_path, sheet_name=sheet_name, index=False, engine=engine)
                preview["saved_to"] = output_path

            return preview
        except Exception as e:
            return {"error": str(e)}

    # ==================== 新增：添加列 ====================
    @classmethod
    def add_column(
            cls,
            file_path: str,
            column_name: str,
            expression: str,
            sheet_name: Optional[str] = None,
            save: bool = False,
            output_path: Optional[str] = None,
            engine: str = "openpyxl"
    ) -> Dict[str, Any]:
        """
        添加新列，列值由表达式计算得出。

        Args:
            column_name: 新列名
            expression: 计算表达式，如 "df['A'] + df['B']" 或 "np.sqrt(df['C'])"
        """
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=100000, engine=engine)
            if sheet_name is None:
                sheet_name = 0

            # 安全执行表达式
            allowed_locals = {"df": df, "pd": pd, "np": np}
            new_vals = eval(expression, {"__builtins__": {}}, allowed_locals)
            df[column_name] = new_vals

            preview = cls._prepare_llm_response(df.head(10), "head", 10, 2000)
            preview["action"] = "add_column"
            preview["column_name"] = column_name
            preview["expression"] = expression
            preview["new_shape"] = df.shape

            if save:
                output_path = output_path or file_path
                df.to_excel(output_path, sheet_name=sheet_name, index=False, engine=engine)
                preview["saved_to"] = output_path

            return preview
        except Exception as e:
            return {"error": str(e)}

    # ==================== 新增：更新单元格 ====================
    @classmethod
    def update_cell(
            cls,
            file_path: str,
            row_index: int,
            column: str,
            value: Any,
            sheet_name: Optional[str] = None,
            save: bool = False,
            output_path: Optional[str] = None,
            engine: str = "openpyxl"
    ) -> Dict[str, Any]:
        """
        更新指定单元格。

        Args:
            row_index: 行索引（0-based，不含表头）
            column: 列名
            value: 新值
        """
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=100000, engine=engine)
            if sheet_name is None:
                sheet_name = 0

            if row_index < 0 or row_index >= len(df):
                return {"error": f"行索引 {row_index} 超出范围 (0-{len(df) - 1})"}
            if column not in df.columns:
                return {"error": f"列 '{column}' 不存在"}

            old_value = df.at[row_index, column]
            df.at[row_index, column] = value

            preview = cls._prepare_llm_response(df.head(10), "head", 10, 2000)
            preview["action"] = "update_cell"
            preview["row_index"] = row_index
            preview["column"] = column
            preview["old_value"] = str(old_value)
            preview["new_value"] = str(value)

            if save:
                output_path = output_path or file_path
                df.to_excel(output_path, sheet_name=sheet_name, index=False, engine=engine)
                preview["saved_to"] = output_path

            return preview
        except Exception as e:
            return {"error": str(e)}

    # ==================== 新增：更新整列 ====================
    @classmethod
    def update_column(
            cls,
            file_path: str,
            column: str,
            expression: str,
            sheet_name: Optional[str] = None,
            save: bool = False,
            output_path: Optional[str] = None,
            engine: str = "openpyxl"
    ) -> Dict[str, Any]:
        """
        更新整列的值，由表达式计算得出。

        Args:
            column: 要更新的列名
            expression: 计算表达式，如 "df['A'] * 2" 或 "df['B'].fillna(0)"
        """
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=100000, engine=engine)
            if sheet_name is None:
                sheet_name = 0
            if column not in df.columns:
                return {"error": f"列 '{column}' 不存在"}

            allowed_locals = {"df": df, "pd": pd, "np": np}
            new_vals = eval(expression, {"__builtins__": {}}, allowed_locals)
            df[column] = new_vals

            preview = cls._prepare_llm_response(df.head(10), "head", 10, 2000)
            preview["action"] = "update_column"
            preview["column"] = column
            preview["expression"] = expression

            if save:
                output_path = output_path or file_path
                df.to_excel(output_path, sheet_name=sheet_name, index=False, engine=engine)
                preview["saved_to"] = output_path

            return preview
        except Exception as e:
            return {"error": str(e)}

    # ==================== 新增：删除行 ====================
    @classmethod
    def delete_rows(
            cls,
            file_path: str,
            condition: str,
            sheet_name: Optional[str] = None,
            save: bool = False,
            output_path: Optional[str] = None,
            engine: str = "openpyxl"
    ) -> Dict[str, Any]:
        """
        删除满足条件的行。

        Args:
            condition: pandas 布尔表达式，如 "df['age'] > 100" 或 "df['city'] == 'New York'"
        """
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=100000, engine=engine)
            if sheet_name is None:
                sheet_name = 0

            allowed_locals = {"df": df, "pd": pd, "np": np}
            mask = eval(condition, {"__builtins__": {}}, allowed_locals)
            if not isinstance(mask, pd.Series) or mask.dtype != bool:
                return {"error": "条件表达式必须返回布尔型 Series"}

            deleted_count = mask.sum()
            df = df[~mask]

            preview = cls._prepare_llm_response(df.head(10), "head", 10, 2000)
            preview["action"] = "delete_rows"
            preview["condition"] = condition
            preview["deleted_count"] = int(deleted_count)
            preview["remaining_rows"] = len(df)

            if save:
                output_path = output_path or file_path
                df.to_excel(output_path, sheet_name=sheet_name, index=False, engine=engine)
                preview["saved_to"] = output_path

            return preview
        except Exception as e:
            return {"error": str(e)}

    # ==================== 新增：删除列 ====================
    @classmethod
    def delete_columns(
            cls,
            file_path: str,
            columns: List[str],
            sheet_name: Optional[str] = None,
            save: bool = False,
            output_path: Optional[str] = None,
            engine: str = "openpyxl"
    ) -> Dict[str, Any]:
        """
        删除指定列。

        Args:
            columns: 要删除的列名列表
        """
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=100000, engine=engine)
            if sheet_name is None:
                sheet_name = 0

            existing_cols = [c for c in columns if c in df.columns]
            if not existing_cols:
                return {"error": "指定的列都不存在"}

            df = df.drop(columns=existing_cols)

            preview = cls._prepare_llm_response(df.head(10), "head", 10, 2000)
            preview["action"] = "delete_columns"
            preview["deleted_columns"] = existing_cols
            preview["new_columns"] = df.columns.tolist()

            if save:
                output_path = output_path or file_path
                df.to_excel(output_path, sheet_name=sheet_name, index=False, engine=engine)
                preview["saved_to"] = output_path

            return preview
        except Exception as e:
            return {"error": str(e)}

    # ==================== 新增：应用函数 ====================
    @classmethod
    def apply_function(
            cls,
            file_path: str,
            func_expr: str,
            axis: int = 0,  # 0: 列, 1: 行
            sheet_name: Optional[str] = None,
            return_preview: bool = True,
            token_budget: int = 4000
    ) -> Dict[str, Any]:
        """
        对数据应用自定义函数，返回结果（聚合或转换）。

        Args:
            func_expr: 函数表达式，如 "lambda x: x.sum()" 或 "lambda x: x.max() - x.min()"
            axis: 0 按列，1 按行
            return_preview: 是否返回结果预览（若为 False，仅返回结果值）
        """
        try:
            # 读取全部数据（限制 10 万行）
            df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=100000)

            allowed_locals = {"df": df, "pd": pd, "np": np}
            func = eval(func_expr, {"__builtins__": {}}, allowed_locals)
            if not callable(func):
                return {"error": "表达式必须返回可调用对象（函数）"}

            result = df.apply(func, axis=axis)

            if return_preview:
                if isinstance(result, pd.Series):
                    preview_data = result.to_dict()
                elif isinstance(result, pd.DataFrame):
                    preview_data = cls._df_to_markdown(result.head(20))
                else:
                    preview_data = str(result)

                return {
                    "success": True,
                    "result": preview_data,
                    "type": str(type(result)),
                    "shape": result.shape if hasattr(result, 'shape') else None
                }
            else:
                return {"success": True, "result": result}
        except Exception as e:
            return {"error": str(e)}

    # ==================== 新增：分组聚合 ====================
    @classmethod
    def aggregate(
            cls,
            file_path: str,
            group_by: Union[str, List[str]],
            agg_dict: Dict[str, Union[str, List[str]]],
            sheet_name: Optional[str] = None,
            token_budget: int = 4000
    ) -> Dict[str, Any]:
        """
        分组聚合，支持大文件分块处理。

        Args:
            group_by: 分组列名（字符串或列表）
            agg_dict: 聚合字典，如 {'sales': 'sum', 'profit': ['mean', 'max']}
        """
        try:
            # 使用分块读取并聚合（支持大文件）
            chunk_size = 50000
            aggregator = {}

            # 初始化聚合器
            def init_agg(col, aggs):
                for agg in aggs if isinstance(aggs, list) else [aggs]:
                    if agg == 'sum':
                        aggregator[(col, 'sum')] = {}
                    elif agg == 'mean':
                        aggregator[(col, 'mean')] = {'sum': 0, 'count': 0}
                    elif agg == 'count':
                        aggregator[(col, 'count')] = {}
                    elif agg == 'min':
                        aggregator[(col, 'min')] = {}
                    elif agg == 'max':
                        aggregator[(col, 'max')] = {}
                    # 可扩展其他聚合

            # 第一遍：初始化聚合器结构
            df_sample = pd.read_excel(file_path, sheet_name=sheet_name, nrows=1)
            if sheet_name is None:
                sheet_name = 0
            for col, aggs in agg_dict.items():
                if col not in df_sample.columns:
                    return {"error": f"列 '{col}' 不存在"}
                init_agg(col, aggs)

            # 分块读取并聚合
            for chunk in pd.read_excel(
                    file_path,
                    sheet_name=sheet_name,
                    chunksize=chunk_size,
                    usecols=list(set([group_by] if isinstance(group_by, str) else group_by) | set(agg_dict.keys()))
            ):
                grouped = chunk.groupby(group_by)
                for col, aggs in agg_dict.items():
                    for agg in aggs if isinstance(aggs, list) else [aggs]:
                        key = (col, agg)
                        if agg == 'sum':
                            series = grouped[col].sum()
                            for idx, val in series.items():
                                aggregator[key][idx] = aggregator[key].get(idx, 0) + val
                        elif agg == 'mean':
                            series_sum = grouped[col].sum()
                            series_count = grouped[col].count()
                            for idx in series_sum.index:
                                if idx not in aggregator[key]:
                                    aggregator[key][idx] = {'sum': 0, 'count': 0}
                                aggregator[key][idx]['sum'] += series_sum[idx]
                                aggregator[key][idx]['count'] += series_count[idx]
                        elif agg == 'count':
                            series = grouped[col].count()
                            for idx, val in series.items():
                                aggregator[key][idx] = aggregator[key].get(idx, 0) + val
                        elif agg == 'min':
                            series = grouped[col].min()
                            for idx, val in series.items():
                                if idx not in aggregator[key] or val < aggregator[key][idx]:
                                    aggregator[key][idx] = val
                        elif agg == 'max':
                            series = grouped[col].max()
                            for idx, val in series.items():
                                if idx not in aggregator[key] or val > aggregator[key][idx]:
                                    aggregator[key][idx] = val

            # 构建结果 DataFrame
            if not aggregator:
                return {"error": "无聚合结果"}

            # 获取所有分组索引
            all_indices = set()
            for key, dict_val in aggregator.items():
                all_indices.update(dict_val.keys())
            all_indices = sorted(all_indices)

            # 构建多级列索引
            result_dict = {idx: {} for idx in all_indices}
            for (col, agg), dict_val in aggregator.items():
                for idx, val in dict_val.items():
                    if agg == 'mean':
                        val = val['sum'] / val['count'] if val['count'] > 0 else None
                    result_dict[idx][(col, agg)] = val

            # 转为 DataFrame
            result_df = pd.DataFrame.from_dict(result_dict, orient='index')
            result_df.columns = pd.MultiIndex.from_tuples(result_df.columns)

            # 生成预览
            preview = cls._prepare_llm_response(result_df.head(20), "head", 20, token_budget)
            preview["success"] = True
            preview["group_by"] = group_by
            preview["aggregations"] = agg_dict

            return preview
        except Exception as e:
            return {"error": str(e)}

    # ==================== 新增：详细统计描述 ====================
    @classmethod
    def describe(
            cls,
            file_path: str,
            columns: Optional[List[str]] = None,
            sheet_name: Optional[str] = None,
            percentiles: Optional[List[float]] = None,
            token_budget: int = 4000
    ) -> Dict[str, Any]:
        """
        生成详细的统计描述（类似 pandas describe），支持大文件分块计算。

        Args:
            columns: 要描述的列名列表，默认所有数值列
            percentiles: 百分位数列表，默认 [0.25, 0.5, 0.75]
        """
        try:
            percentiles = percentiles or [0.25, 0.5, 0.75]
            # 先读取少量数据获取列信息
            sample_df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=5)
            if sheet_name is None:
                sheet_name = 0

            # 确定要描述的列
            if columns is None:
                cols = sample_df.select_dtypes(include=[np.number]).columns.tolist()
            else:
                cols = [c for c in columns if c in sample_df.columns]

            if not cols:
                return {"error": "没有可描述的数值列"}

            # 分块计算统计量
            stats = {col: {'count': 0, 'sum': 0, 'sum2': 0, 'min': float('inf'), 'max': float('-inf')} for col in cols}
            # 百分位数需要分位数，分块计算较复杂，采用近似：先收集样本分位数（可读全部数据，但限制行数）
            # 简单起见，如果文件不大（<500k行），直接读取；否则采样
            total_rows = cls._detect_metadata(file_path, sheet_name, 'openpyxl').get('total_rows', 0)
            if total_rows < 500000:
                df = pd.read_excel(file_path, sheet_name=sheet_name, usecols=cols)
            else:
                # 采样 10% 或 10000 行
                n_sample = min(10000, int(total_rows * 0.1))
                df = pd.read_excel(file_path, sheet_name=sheet_name, usecols=cols, nrows=n_sample)

            # 计算描述
            desc = df.describe(percentiles=percentiles)
            # 转为 Markdown
            md = cls._df_to_markdown(desc)

            return {
                "success": True,
                "description": md,
                "columns_described": cols,
                "sample_size": len(df),
                "total_rows_estimate": total_rows,
                "token_estimate": len(md) // 4
            }
        except Exception as e:
            return {"error": str(e)}

    # ==================== 新增：相关性矩阵 ====================
    @classmethod
    def correlation(
            cls,
            file_path: str,
            columns: Optional[List[str]] = None,
            sheet_name: Optional[str] = None,
            method: str = 'pearson',
            token_budget: int = 4000
    ) -> Dict[str, Any]:
        """
        计算数值列之间的相关系数矩阵。

        Args:
            method: 'pearson', 'kendall', 'spearman'
        """
        try:
            sample_df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=5)
            if sheet_name is None:
                sheet_name = 0

            if columns is None:
                cols = sample_df.select_dtypes(include=[np.number]).columns.tolist()
            else:
                cols = [c for c in columns if c in sample_df.columns]

            if len(cols) < 2:
                return {"error": "至少需要两列数值列"}

            # 读取数据（限制行数，或采样）
            total_rows = cls._detect_metadata(file_path, sheet_name, 'openpyxl').get('total_rows', 0)
            if total_rows < 500000:
                df = pd.read_excel(file_path, sheet_name=sheet_name, usecols=cols)
            else:
                n_sample = min(10000, int(total_rows * 0.1))
                df = pd.read_excel(file_path, sheet_name=sheet_name, usecols=cols, nrows=n_sample)

            corr = df.corr(method=method)
            md = cls._df_to_markdown(corr)

            return {
                "success": True,
                "correlation_matrix": md,
                "method": method,
                "columns": cols,
                "sample_size": len(df)
            }
        except Exception as e:
            return {"error": str(e)}
